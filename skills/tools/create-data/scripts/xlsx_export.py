"""XLSX export with randomized styling and layout variations.

Variations per export:
- Color theme: 5 distinct palettes (blue, green, gold, red, purple)
- Sheet layout: single sheet vs multi-sheet grouped by request_type
- Header font size: 10, 11, or 12pt
- Optional: filters, freeze panes, cell borders, alternating row colors
- Optional: summary statistics sheet with totals and quality metrics
"""

from __future__ import annotations

import json
import random
from pathlib import Path
from typing import Any


def export_to_xlsx(data: list[dict[str, Any]], path: str | Path) -> Path:
    """Export data to XLSX with randomized styling, layout, and sheet structure."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    dest = Path(path).with_suffix(".xlsx")
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    color_themes = [
        {"header_bg": "1F4E79", "header_fg": "FFFFFF", "alt_row": "D6E4F0"},
        {"header_bg": "548235", "header_fg": "FFFFFF", "alt_row": "E2EFDA"},
        {"header_bg": "BF8F00", "header_fg": "FFFFFF", "alt_row": "FFF2CC"},
        {"header_bg": "C00000", "header_fg": "FFFFFF", "alt_row": "FCE4EC"},
        {"header_bg": "7030A0", "header_fg": "FFFFFF", "alt_row": "E8D5F5"},
    ]
    theme = random.choice(color_themes)
    use_multi_sheet = random.choice([True, False]) and len(data) > 3
    use_filters = random.choice([True, False])
    use_freeze = random.choice([True, False])
    use_borders = random.choice([True, False])
    use_alt_rows = random.choice([True, False])
    include_summary = random.choice([True, False])

    header_font = Font(bold=True, color=theme["header_fg"], size=random.choice([10, 11, 12]))
    header_fill = PatternFill(start_color=theme["header_bg"], end_color=theme["header_bg"], fill_type="solid")
    alt_fill = PatternFill(start_color=theme["alt_row"], end_color=theme["alt_row"], fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")

    columns = ["input", "output"]
    meta_keys_set: list[str] = []
    for entry in data:
        for k in (entry.get("metadata") or {}):
            if k not in meta_keys_set:
                meta_keys_set.append(k)
    columns.extend(meta_keys_set)
    if any(e.get("transformation_rules") for e in data):
        columns.insert(0, "transformation_rules")

    def _write_sheet(ws, entries: list[dict[str, Any]], cols: list[str]):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_align
            if use_borders:
                cell.border = thin_border

        for ri, entry in enumerate(entries, 2):
            for ci, col in enumerate(cols, 1):
                if col == "transformation_rules":
                    val = "; ".join(entry.get("transformation_rules", []))
                elif col in ("input", "output"):
                    val = entry.get(col, "")
                else:
                    meta = entry.get("metadata") or {}
                    v = meta.get(col, "")
                    val = json.dumps(v) if isinstance(v, (list, dict)) else str(v)

                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = wrap_align
                if use_borders:
                    cell.border = thin_border
                if use_alt_rows and ri % 2 == 0:
                    cell.fill = alt_fill

        for ci in range(1, len(cols) + 1):
            col_letter = get_column_letter(ci)
            ws.column_dimensions[col_letter].width = random.choice([20, 30, 40, 50])

        if use_freeze:
            ws.freeze_panes = "A2"
        if use_filters:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(entries) + 1}"

    if use_multi_sheet:
        groups: dict[str, list[dict[str, Any]]] = {}
        for entry in data:
            key = (entry.get("metadata") or {}).get("request_type", "Other")
            groups.setdefault(key, []).append(entry)

        first = True
        for group_name, group_entries in groups.items():
            if first:
                ws = wb.active
                ws.title = group_name[:31]
                first = False
            else:
                ws = wb.create_sheet(title=group_name[:31])
            _write_sheet(ws, group_entries, columns)
    else:
        ws = wb.active
        ws.title = "Data"
        _write_sheet(ws, data, columns)

    if include_summary:
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.cell(row=1, column=1, value="Metric").font = header_font
        ws_sum.cell(row=1, column=2, value="Value").font = header_font
        ws_sum.cell(row=1, column=1).fill = header_fill
        ws_sum.cell(row=1, column=2).fill = header_fill

        stats = [
            ("Total Entries", len(data)),
            ("Unique Request Types", len({(e.get("metadata") or {}).get("request_type", "") for e in data})),
        ]
        scores = [
            (e.get("metadata") or {}).get("quality_score")
            for e in data
            if (e.get("metadata") or {}).get("quality_score") is not None
        ]
        if scores:
            stats.extend([
                ("Avg Quality Score", f"{sum(scores) / len(scores):.1f}"),
                ("Min Quality Score", min(scores)),
                ("Max Quality Score", max(scores)),
            ])

        for ri, (metric, value) in enumerate(stats, 2):
            ws_sum.cell(row=ri, column=1, value=metric)
            ws_sum.cell(row=ri, column=2, value=value)

        ws_sum.column_dimensions["A"].width = 25
        ws_sum.column_dimensions["B"].width = 20

    wb.save(dest)
    return dest
